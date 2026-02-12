from datetime import datetime, timedelta
import json
import os
from openpyxl import load_workbook
from lector_excel import buscar_cedula
from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse

# 🚀 Flask app
app = Flask(__name__)

# 🧩 Sesiones por usuario (multiusuario)
sesiones = {}

# 📘 Cargar menú desde JSON
def cargar_menu():
    with open("menu.json", "r", encoding="utf-8") as f:
        return json.load(f)

menu = cargar_menu()

# 📁 Carpeta de archivos TXT
RUTA_TXT = "txt"

# 🔹 Mostrar menú principal
def mostrar_menu_principal():
    texto = "\n📋 *MENÚ PRINCIPAL*\n"
    for clave, item in menu.items():
        texto += f"{clave}. {item['titulo']}\n"
    texto += "\n➡️ Responde con el número de tu elección:"
    return texto

# 🔹 Mostrar submenú
def mostrar_submenu(opcion):
    sub = menu[opcion]["subopciones"]
    texto = f"\n📂 *{menu[opcion]['titulo']}*\n"
    for clave, item in sub.items():
        texto += f"{clave}. {item}\n"
    texto += "\n⬅️ Escribe 0 para volver al menú principal."
    return texto

# 🔹 Leer archivo TXT
def leer_txt(nombre_archivo):
    ruta = os.path.join(RUTA_TXT, f"{nombre_archivo}.txt")
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "❌ Archivo de información no encontrado."
    
def limpiar_sesiones():
    ahora = datetime.now()
    for uid in list(sesiones.keys()):
        ultimo = sesiones[uid].get("ultimo")
        if ultimo and (ahora - ultimo > timedelta(minutes=30)):
            del sesiones[uid]

# 🔹 Funciones Excel (copiadas de tu código funcional)
def obtener_horario(usuario):
    archivo = os.path.join("datos", usuario.get("curso", "").strip() + ".xlsx")
    if not os.path.exists(archivo):
        return f"❌ No se encontró el archivo del curso: {usuario.get('curso', '')}"
    try:
        wb = load_workbook(filename=archivo, data_only=True)
        if "Horario" not in wb.sheetnames:
            return "❌ Hoja 'Horario' no encontrada."
        ws = wb["Horario"]
        contenido = ""
        for row in ws.iter_rows(values_only=True):
            fila = [str(celda) for celda in row if celda]
            if fila:
                contenido += " | ".join(fila) + "\n"
        return f"🕒 *Horario del curso {usuario['curso']}*\n{contenido}" if contenido else "❌ No se encontró horario para este curso."
    except Exception as e:
        return f"❌ Error al obtener horario: {str(e)}"

def obtener_horario_docente(usuario):
    try:
        archivo = os.path.join("datos", "docentes.xlsx")
        if not os.path.exists(archivo):
            return "❌ Archivo de docentes no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Horario" not in wb.sheetnames:
            return "❌ Hoja 'Horario' no encontrada."
        ws = wb["Horario"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cedula_excel, link = row
            if str(cedula_excel).strip() == str(usuario["cedula"]).strip():
                return f"🕒 *Horario del Docente*\n{link}"
        return "❌ No se encontró horario asignado para tu cédula."
    except Exception as e:
        return f"❌ Error al obtener horario: {str(e)}"

def obtener_materias_docente(usuario):
    try:
        archivo = os.path.join("datos", "docentes.xlsx")
        if not os.path.exists(archivo):
            return "❌ Archivo de docentes no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Materias" not in wb.sheetnames:
            return "❌ Hoja 'Materias' no encontrada."
        ws = wb["Materias"]
        materias = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cedula_excel, materia = row
            if str(cedula_excel).strip() == str(usuario["cedula"]).strip() and materia:
                materias.append(str(materia))
        return "📚 *Materias que dictas:*\n- " + "\n- ".join(materias) if materias else "❌ No se encontraron materias asignadas a tu cédula."
    except Exception as e:
        return f"❌ Error al obtener materias: {str(e)}"

def obtener_claves(usuario):
    try:
        archivo = os.path.join("datos", usuario.get("archivo", ""))
        if not os.path.exists(archivo):
            return "❌ Archivo del curso no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Claves" not in wb.sheetnames:
            return "❌ Hoja 'Claves' no encontrada."
        ws = wb["Claves"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cedula_excel, *resto = row
            if str(cedula_excel).strip() == str(usuario["cedula"]).strip():
                if len(resto) == 1:
                    contraseña = resto[0]
                    return f"🔐 *Acceso a la plataforma educativa*\n👤 Cédula: {cedula_excel}\n🔑 Contraseña: {contraseña}"
                elif len(resto) >= 2:
                    usuario_plat, contraseña = resto[:2]
                    return f"🔐 *Acceso a la plataforma educativa*\n👤 Usuario: {usuario_plat}\n🔑 Contraseña: {contraseña}"
        return "❌ No se encontraron credenciales para esta cédula."
    except Exception as e:
        return f"❌ Error al obtener las claves: {str(e)}"

def obtener_materias(usuario):
    try:
        archivo = os.path.join("datos", usuario.get("curso", "").strip() + ".xlsx")
        if not os.path.exists(archivo):
            return "❌ Archivo del curso no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Materias" not in wb.sheetnames:
            return "❌ Hoja 'Materias' no encontrada."
        ws = wb["Materias"]
        materias = [str(row[0]) for row in ws.iter_rows(values_only=True) if row[0]]
        return "📚 *Materias del curso {}*:\n- ".format(usuario["curso"]) + "\n- ".join(materias) if materias else "❌ No se encontraron materias."
    except Exception as e:
        return f"❌ Error al obtener materias: {str(e)}"

def obtener_profesores(usuario):
    try:
        archivo = os.path.join("datos", usuario.get("curso", "").strip() + ".xlsx")
        if not os.path.exists(archivo):
            return "❌ Archivo del curso no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Profesores" not in wb.sheetnames:
            return "❌ Hoja 'Profesores' no encontrada."
        ws = wb["Profesores"]
        profesores = [str(row[0]) for row in ws.iter_rows(values_only=True) if row[0]]
        return "👨‍🏫 *Profesores del curso {}*:\n- ".format(usuario["curso"]) + "\n- ".join(profesores) if profesores else "❌ No se encontraron profesores."
    except Exception as e:
        return f"❌ Error al obtener profesores: {str(e)}"

def obtener_valores_pendientes(usuario):
    try:
        archivo = os.path.join("datos", usuario.get("archivo", ""))
        if not os.path.exists(archivo):
            return "❌ Archivo del curso no encontrado."
        wb = load_workbook(filename=archivo, data_only=True)
        if "Pagos" not in wb.sheetnames:
            return "❌ Hoja 'Pagos' no encontrada."
        ws = wb["Pagos"]
        pendientes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cedula_excel, mes, monto = row
            if str(cedula_excel).strip() == str(usuario["cedula"]).strip():
                pendientes.append((mes, monto))
        if not pendientes:
            return "✅ No tienes valores pendientes."
        mensaje = f"💰 *Valores pendientes para {usuario['nombre']}*:\n"
        for mes, monto in pendientes:
            mensaje += f"- {mes}: ${monto}\n"
        return mensaje
    except Exception as e:
        return f"❌ Error al obtener valores pendientes: {str(e)}"

def procesar_mensaje_multiusuario(mensaje, sesion):
    mensaje = mensaje.strip()
    ahora = datetime.now()

    # 🚪 Salir del chatbot en cualquier momento
    if mensaje in ["salir", "exit", "cancelar"]:
        sesion.update({
            "usuario": {"rol": None, "nombre": None, "curso": None, "archivo": None, "cedula": None},
            "nivel": "menu_principal",
            "opcion": None,
            "ultimo": ahora
        })
        return (
            "🔄 Has salido del chatbot.\n\n"
            "👋 ¡Hola! Soy *Lukibot*.\n"
            "Por favor ingresa tu número de cédula para iniciar nuevamente."
        )

    usuario_actual = sesion["usuario"]
    nivel_actual = sesion["nivel"]
    opcion_actual = sesion["opcion"]
    ultimo_mensaje = sesion.get("ultimo")

    # ⏰ Expiración por inactividad
    if ultimo_mensaje and (ahora - ultimo_mensaje > timedelta(minutes=10)):
        sesion.update({
            "usuario": {"rol": None, "nombre": None, "curso": None, "archivo": None, "cedula": None},
            "nivel": "menu_principal",
            "opcion": None,
            "ultimo": ahora
        })
        return ("⏰ La sesión se cerró por inactividad.\n\n"
                "👋 ¡Hola! Soy *Lukibot*, el asistente virtual de la *Unidad Educativa María Luisa Luque de Sotomayor*.\n"
                "Por favor ingresa tu número de cédula, si eres docente ingresa tu usuario o contraseña.")

    sesion["ultimo"] = ahora

 # 🔐 Inicio / cédula
    if usuario_actual["rol"] is None:

            info = buscar_cedula(mensaje)

            if info:
                info["archivo"] = info.get("curso", "").strip() + ".xlsx"
                info["cedula"] = mensaje
                sesion["usuario"] = info
                rol = info["rol"].upper()
                sesion["nivel"] = "menu_principal"
                return f"✅ Bienvenido {info['nombre']}. Has ingresado como *{rol}*.\n" + mostrar_menu_principal()
            else:
                return ("👋 ¡Hola! Soy *Lukibot* 🤖\n\n"
                "🔐 Ingresa tu usuario (docentes) o tu número de cédula (estudiantes/padres).")

    # 📋 Menú principal
    if nivel_actual == "menu_principal":
        if mensaje in menu:
            sesion["opcion"] = mensaje
            sesion["nivel"] = "submenu"
            return mostrar_submenu(mensaje)
        else:
            return "⚠ Opción no válida."

    # 📂 Submenú
    if nivel_actual == "submenu":
        if mensaje == "0":
            sesion["nivel"] = "menu_principal"
            return mostrar_menu_principal()

        sub = menu[opcion_actual]["subopciones"]
        if mensaje in sub:
            opcion_texto = sub[mensaje]

            # ⚠ Restricciones para estudiantes
            if usuario_actual["rol"] == "estudiante" and opcion_texto in [
                "Solicitar claves del Wi-Fi institucional",
                "Reglamento interno para docentes"
            ]:
                return "🚫 No tienes permiso para acceder a esta opción."

            # 🔹 Manejo de "Salir del chatbot" opción 10
            if opcion_actual == "10":  # Opción salir
                if mensaje == "1" or opcion_texto.lower() == "finalizar conversación":
                    sesion.update({
                        "usuario": {"rol": None, "nombre": None, "curso": None, "archivo": None, "cedula": None},
                        "nivel": "menu_principal",
                        "opcion": None,
                        "ultimo": ahora
                    })
                    return "🔄 Sesión finalizada. Por favor ingresa tu número de cédula para iniciar nuevamente."
                if mensaje == "2" or opcion_texto.lower() == "volver al inicio":
                    sesion["nivel"] = "menu_principal"
                    sesion["opcion"] = None
                    return mostrar_menu_principal()

            # 🔹 Llamadas automáticas a funciones según texto
            if "horario" in opcion_texto.lower():
                if usuario_actual["rol"] == "docente":
                    return obtener_horario_docente(usuario_actual)
                else:
                    return obtener_horario(usuario_actual)
            if "materias" in opcion_texto.lower():
                if usuario_actual["rol"] == "docente":
                    return obtener_materias_docente(usuario_actual)
                else:
                    return obtener_materias(usuario_actual)
            if "profesores" in opcion_texto.lower():
                return obtener_profesores(usuario_actual)
            if "plataforma educativa" in opcion_texto.lower():
                return obtener_claves(usuario_actual)
            if "valores pendientes" in opcion_texto.lower():
                if usuario_actual["rol"] == "docente":
                    return "🚫 Estimado docente, esta opción no está disponible para su rol."
                return obtener_valores_pendientes(usuario_actual)

            # TXT
            txt = leer_txt(opcion_texto)
            if txt != "❌ Archivo de información no encontrado.":
                return txt

            return f"📄 Has seleccionado: *{opcion_texto}*"
        else:
            return "⚠ Opción no válida."

    return "❓ No entendí tu mensaje."

# 🔹 Webhook Flask
@app.route("/webhook", methods=["POST"])
def webhook():
    limpiar_sesiones()
    
    mensaje = request.form.get("Body", "").strip().lower()
    usuario_id = request.form.get("From")

    # Crear sesión si no existe
    if usuario_id not in sesiones:
        sesiones[usuario_id] = {
            "usuario": {"rol": None, "nombre": None, "curso": None, "archivo": None, "cedula": None},
            "nivel": "menu_principal",
            "opcion": None,
            "ultimo": None
        }

    sesion = sesiones[usuario_id]
    respuesta = procesar_mensaje_multiusuario(mensaje, sesion)

    resp = MessagingResponse()
    resp.message(respuesta)
    return str(resp)

@app.route("/", methods=["GET"])
def home():
    return "Servidor Flask activo ✅"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)

